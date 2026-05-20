"""
KAM Supply Intelligence Agent — agent_format_answer.py
Sprint 2, US-09: format_answer (Node 6) — now live

Changes from agent_SQL.py (US-08):
  - Node 6 (format_answer) is now fully implemented:
      · Structured template — no extra LLM call, deterministic output
      · Three section layout: Client Profile / Operational Data / Transparency
      · Question-type-aware formatting:
          supplier_count  → count summary line
          product_list    → table of rate codes per supplier
          product_details → detailed table with route and rate type
      · Optional CSV export (set export_csv=True in run_agent())
          CSV saved to ./exports/<client>_<question_type>_<date>.csv
          Salesforce profile written as header comment rows
      · Cost footer on every answer

Graph flow — ALL NODES LIVE:
  understand_question       ✅ (US-05)
        ↓
  fetch_salesforce_client   ✅ (US-06)
        ↓
  retrieve_schema           ✅ (US-07)
        ↓
  generate_sql              ✅ (US-08)
        ↓
  execute_sql               ✅ (US-08)
        |── (error, retry_count < 2) ──→ generate_sql
        ↓ (success)
  format_answer             ✅ (US-09)  ← new
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
from datetime import date
from typing import TypedDict, Optional, List
from dotenv import load_dotenv

import requests
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage
from langgraph.graph import StateGraph, END

os.environ["ANONYMIZED_TELEMETRY"] = "False"
load_dotenv()

# ── LLM ──────────────────────────────────────────────────────────────────────

llm = ChatOpenAI(
    model="gpt-4o-mini",
    temperature=0,
    openai_api_key=os.getenv("OPENAI_API_KEY"),
)

# ── ChromaDB config ───────────────────────────────────────────────────────────

CHROMA_PATH       = "./chroma_db"
CHROMA_COLLECTION = "kam_schema_store"
CHROMA_N_RESULTS  = 3

# ── Pricing ───────────────────────────────────────────────────────────────────

PRICING = {
    "gpt-4o-mini": {
        "prompt":     0.000150 / 1000,
        "completion": 0.000600 / 1000,
    },
    "text-embedding-3-small": {
        "prompt":     0.000020 / 1000,
        "completion": 0.0,
    },
}
SUPABASE_COST_PER_QUERY = 0.0

# ── CSV export folder ─────────────────────────────────────────────────────────

EXPORT_DIR = "./exports"


# ── Agent state ───────────────────────────────────────────────────────────────

class AgentState(TypedDict):
    # Input
    question: str
    export_csv: bool             # if True, write CSV file after format_answer

    # Node 1
    client_name: Optional[str]
    question_type: Optional[str]
    intent_summary: Optional[str]

    # Node 2
    salesforce_data: Optional[dict]
    salesforce_error: Optional[str]

    # Node 3
    schema_context: Optional[str]
    schema_doc_ids: Optional[List[str]]

    # Node 4
    sql_query: Optional[str]

    # Node 5
    sql_result: Optional[list]
    sql_error: Optional[str]
    retry_count: int

    # Node 6
    final_answer: Optional[str]
    csv_path: Optional[str]      # set if CSV was exported

    # Cost tracking
    usage: List[dict]
    cost_summary: Optional[dict]


# ── Node 1: understand_question ───────────────────────────────────────────────

UNDERSTAND_SYSTEM_PROMPT = """You are a parser for a KAM Supply Intelligence Agent
in the car rental distribution industry.

Your job: extract structured information from a Key Account Manager's question.

CLIENTS in the system: Check24, Autoslash, HappyCar

QUESTION TYPES:
- supplier_count   → "how many suppliers does [client] have?"
- supplier_list    → "which/what suppliers does [client] have?"
- product_list     → "which/what products does [supplier] have for [client]?"
- product_details  → "details/specifics about products, routes, rates, excess for [client]"

Respond ONLY with a JSON object — no markdown, no explanation:
{
  "client_name": "<name exactly as it appears in the system, or null if not found>",
  "question_type": "<supplier_count | supplier_list | product_list | product_details>",
  "intent_summary": "<one sentence describing what the KAM wants to know>"
}"""


def understand_question(state: AgentState) -> AgentState:
    print(f"\n[Node 1] understand_question")
    print(f"  Question: {state['question']}")

    messages = [
        SystemMessage(content=UNDERSTAND_SYSTEM_PROMPT),
        HumanMessage(content=state["question"]),
    ]
    response      = llm.invoke(messages)
    raw           = response.content.strip()

    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    parsed         = json.loads(raw)
    client_name    = parsed.get("client_name")
    question_type  = parsed.get("question_type")
    intent_summary = parsed.get("intent_summary")

    print(f"  → client_name:    {client_name}")
    print(f"  → question_type:  {question_type}")
    print(f"  → intent_summary: {intent_summary}")

    usage_entry = {
        "node":              "understand_question",
        "model":             "gpt-4o-mini",
        "prompt_tokens":     response.usage_metadata.get("input_tokens", 0),
        "completion_tokens": response.usage_metadata.get("output_tokens", 0),
    }
    print(f"  → tokens: {usage_entry['prompt_tokens']} in / {usage_entry['completion_tokens']} out")

    if not client_name:
        print(f"  WARNING: Client not recognised — short-circuiting to END")
        return {
            **state,
            "client_name":    None,
            "question_type":  question_type,
            "intent_summary": intent_summary,
            "usage":          state.get("usage", []) + [usage_entry],
            "final_answer": (
                "I couldn't identify the client in your question.\n"
                "The clients I currently support are: *Check24*, *Autoslash*, and *HappyCar*.\n"
                "Please check the name and try again."
            ),
        }

    return {
        **state,
        "client_name":    client_name,
        "question_type":  question_type,
        "intent_summary": intent_summary,
        "usage":          state.get("usage", []) + [usage_entry],
    }


# ── Salesforce helpers ────────────────────────────────────────────────────────

SF_TOKEN_EXPIRED_CODES = {"INVALID_SESSION_ID", "SESSION_EXPIRED"}

class SalesforceAuthError(Exception):
    pass

def _get_sf_token() -> dict:
    payload = {
        "grant_type":    "client_credentials",
        "client_id":     os.getenv("SF_CONSUMER_KEY"),
        "client_secret": os.getenv("SF_CONSUMER_SECRET"),
    }
    r = requests.post(
        f"https://{os.getenv('SF_ORG_DOMAIN')}/services/oauth2/token",
        data=payload,
        timeout=10,
    )
    token_data = r.json()
    if "access_token" not in token_data:
        raise SalesforceAuthError(f"Token request failed: {token_data}")
    return token_data

def _is_token_expired_error(e: Exception) -> bool:
    class_name = type(e).__name__
    msg        = str(e).upper()
    return (
        "EXPIREDSESSION" in class_name.upper()
        or any(code in msg for code in SF_TOKEN_EXPIRED_CODES)
    )

SF_TYPE_MAP = {
    "Channel Partner / Reseller": "Commissionable",
    "Technology Partner":         "Wholesaler",
}
SF_PRIORITY_MAP = {
    "High":   "Strategic",
    "Medium": "Growth",
    "Low":    "Standard",
}


# ── Node 2: fetch_salesforce_client ───────────────────────────────────────────

def fetch_salesforce_client(state: AgentState) -> AgentState:
    from simple_salesforce import Salesforce

    client_name = state.get("client_name")
    print(f"\n[Node 2] fetch_salesforce_client")
    print(f"  Client: {client_name}")

    def _run_soql(token_data: dict) -> dict:
        sf = Salesforce(
            instance_url=token_data["instance_url"],
            session_id=token_data["access_token"],
        )
        soql = (
            f"SELECT Name, AccountNumber, Type, CustomerPriority__c, Active__c, Owner.Name "
            f"FROM Account WHERE Name = '{client_name}' LIMIT 1"
        )
        result = sf.query(soql)
        if result["totalSize"] == 0:
            raise Exception(f"Account '{client_name}' not found in Salesforce")
        raw = result["records"][0]
        return {
            "Name":            raw.get("Name"),
            "account_number":  raw.get("AccountNumber", "N/A"),
            "Type":            raw.get("Type"),
            "business_model":  SF_TYPE_MAP.get(raw.get("Type", ""), raw.get("Type", "N/A")),
            "account_tier":    SF_PRIORITY_MAP.get(raw.get("CustomerPriority__c", ""), raw.get("CustomerPriority__c", "N/A")),
            "contract_status": "Active" if raw.get("Active__c") else "Inactive",
            "kam":             raw.get("Owner", {}).get("Name", "N/A"),
        }

    try:
        token_data = _get_sf_token()
        try:
            salesforce_data = _run_soql(token_data)
        except Exception as query_err:
            if _is_token_expired_error(query_err):
                print(f"  ⚠ Token expired — fetching fresh token and retrying...")
                token_data      = _get_sf_token()
                salesforce_data = _run_soql(token_data)
            else:
                raise

        print(f"  ✓ Found in Salesforce:")
        print(f"    business_model:  {salesforce_data['business_model']}")
        print(f"    account_tier:    {salesforce_data['account_tier']}")
        print(f"    contract_status: {salesforce_data['contract_status']}")
        print(f"    kam:             {salesforce_data['kam']}")

        return {**state, "salesforce_data": salesforce_data}

    except Exception as e:
        print(f"  ✗ Salesforce unavailable: {e}")
        return {
            **state,
            "salesforce_data":  None,
            "salesforce_error": str(e),
        }


# ── Node 3: retrieve_schema ───────────────────────────────────────────────────

def retrieve_schema(state: AgentState) -> AgentState:
    import chromadb
    from chromadb.utils import embedding_functions

    intent = state.get("intent_summary", state.get("question", ""))
    print(f"\n[Node 3] retrieve_schema")
    print(f"  Query: {intent}")

    try:
        chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)
        ef = embedding_functions.OpenAIEmbeddingFunction(
            api_key=os.getenv("OPENAI_API_KEY"),
            model_name="text-embedding-3-small",
        )
        collection = chroma_client.get_collection(
            name=CHROMA_COLLECTION,
            embedding_function=ef,
        )
        results   = collection.query(query_texts=[intent], n_results=CHROMA_N_RESULTS)
        doc_ids   = results["ids"][0]
        documents = results["documents"][0]
        distances = results["distances"][0]

        print(f"  ✓ Retrieved {len(doc_ids)} documents:")
        for doc_id, dist in zip(doc_ids, distances):
            print(f"    {doc_id}  (distance: {dist:.4f})")

        schema_context   = "\n\n---\n\n".join(documents)
        embedding_tokens = max(1, len(intent) // 4)

        usage_entry = {
            "node":              "retrieve_schema",
            "model":             "text-embedding-3-small",
            "prompt_tokens":     embedding_tokens,
            "completion_tokens": 0,
        }
        print(f"  → embedding tokens (est.): {embedding_tokens}")

        return {
            **state,
            "schema_context": schema_context,
            "schema_doc_ids": doc_ids,
            "usage":          state.get("usage", []) + [usage_entry],
        }

    except Exception as e:
        print(f"  ✗ ChromaDB retrieval failed: {e}")
        fallback_context = (
            "Tables: supplier (id, name, code), "
            "client_supplier (id, client_name, supplier_id, status), "
            "product (id, client_name, supplier_id, rate_code, rate_type, "
            "product_type, source_country, destination_country, status)"
        )
        return {
            **state,
            "schema_context": fallback_context,
            "schema_doc_ids": [],
            "usage":          state.get("usage", []) + [{
                "node": "retrieve_schema", "model": "text-embedding-3-small",
                "prompt_tokens": 0, "completion_tokens": 0,
            }],
        }


# ── Node 4: generate_sql ──────────────────────────────────────────────────────

SQL_TABLE_HINTS = """
AVAILABLE TABLES — use these exact names, no others:
  - supplier          (id, name, code, region)
  - client_supplier   (id, client_name, supplier_id, status)
  - product           (id, client_name, supplier_id, rate_code, rate_type,
                       product_type, source_country, destination_country, status, notes)

QUESTION TYPE → SQL PATTERN:
  - supplier_count  → SELECT COUNT(DISTINCT cs.supplier_id) FROM client_supplier cs WHERE ...
  - supplier_list   → SELECT s.name, s.code, s.region FROM supplier s JOIN client_supplier cs ON s.id = cs.supplier_id WHERE cs.client_name = '...' AND cs.status = 'active'
  - product_list    → SELECT rate_code, rate_type, product_type, source_country, destination_country FROM product WHERE client_name = '...' AND status = 'active'
  - product_details → SELECT p.rate_code, p.rate_type, s.name AS supplier, p.source_country, p.destination_country FROM product p JOIN supplier s ON s.id = p.supplier_id WHERE ...

KEY RULES:
  - Always filter by status = 'active' unless the question asks about inactive records
  - client_name values are case-sensitive: 'Check24', 'Autoslash', 'HappyCar'
  - Supplier name values are: 'Avis', 'Hertz', 'Enterprise', 'Budget', 'Sixt'
  - To filter by supplier name always JOIN supplier and use: supplier.name = '<name>'
    NEVER use supplier.code to look up a supplier by name — code is uppercase (e.g. 'AVIS')
    and is NOT the same as the supplier name
  - JOIN supplier using: supplier.id = product.supplier_id
                     or: supplier.id = client_supplier.supplier_id
  - When querying products for a specific client, filter using product.client_name directly
    DO NOT also join client_supplier — this causes duplicate rows
    Only join client_supplier when you need connection-level data (e.g. supplier count per client)
  - Return only a single valid SELECT statement — no explanations, no markdown
"""

GENERATE_SQL_SYSTEM_PROMPT = """You are a SQL generator for a PostgreSQL database
used by a car rental distribution platform.

{table_hints}

SCHEMA CONTEXT (retrieved from knowledge base):
{schema_context}

{retry_block}

Generate a single valid PostgreSQL SELECT statement that answers the question.
- No markdown, no explanation, no backticks — raw SQL only.
- Use only the tables and columns listed above.
- Always alias columns clearly for readability.
"""


def generate_sql(state: AgentState) -> AgentState:
    print(f"\n[Node 4] generate_sql")
    retry = state.get("retry_count", 0)

    retry_block = ""
    if retry > 0:
        sql_error = state.get("sql_error", "unknown error")
        print(f"  Retry #{retry} — previous error: {sql_error}")
        retry_block = f"""PREVIOUS SQL ATTEMPT FAILED:
Error: {sql_error}
Fix the error and generate a corrected SQL query."""

    system_prompt = GENERATE_SQL_SYSTEM_PROMPT.format(
        table_hints    = SQL_TABLE_HINTS,
        schema_context = state.get("schema_context", "No schema context available."),
        retry_block    = retry_block,
    )

    messages = [
        SystemMessage(content=system_prompt),
        HumanMessage(content=state["question"]),
    ]
    response = llm.invoke(messages)
    sql_raw  = response.content.strip()

    if sql_raw.startswith("```"):
        sql_raw = sql_raw.split("```")[1]
        if sql_raw.lower().startswith("sql"):
            sql_raw = sql_raw[3:]
        sql_raw = sql_raw.strip()

    print(f"  → SQL: {sql_raw[:120]}{'...' if len(sql_raw) > 120 else ''}")

    usage_entry = {
        "node":              "generate_sql",
        "model":             "gpt-4o-mini",
        "prompt_tokens":     response.usage_metadata.get("input_tokens", 0),
        "completion_tokens": response.usage_metadata.get("output_tokens", 0),
    }
    print(f"  → tokens: {usage_entry['prompt_tokens']} in / {usage_entry['completion_tokens']} out")

    return {
        **state,
        "sql_query": sql_raw,
        "sql_error": None,
        "usage":     state.get("usage", []) + [usage_entry],
    }


# ── Node 5: execute_sql ───────────────────────────────────────────────────────

def execute_sql(state: AgentState) -> AgentState:
    from supabase import create_client

    sql = state.get("sql_query", "")
    print(f"\n[Node 5] execute_sql")
    print(f"  SQL: {sql[:120]}{'...' if len(sql) > 120 else ''}")

    usage_entry = {
        "node":              "execute_sql",
        "model":             "supabase",
        "prompt_tokens":     0,
        "completion_tokens": 0,
        "supabase_queries":  1,
    }

    try:
        supabase = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_KEY"),
        )
        sql_clean = sql.strip().rstrip(";")
        result = supabase.rpc("execute_sql", {"query": sql_clean}).execute()
        rows   = result.data if result.data else []

        print(f"  ✓ Query returned {len(rows)} row(s)")
        if rows:
            print(f"  → Sample: {json.dumps(rows[0], indent=None)}")

        return {
            **state,
            "sql_result": rows,
            "sql_error":  None,
            "usage":      state.get("usage", []) + [usage_entry],
        }

    except Exception as e:
        error_msg = str(e)
        print(f"  ✗ SQL execution failed: {error_msg}")
        return {
            **state,
            "sql_result":  None,
            "sql_error":   error_msg,
            "retry_count": state.get("retry_count", 0) + 1,
            "usage":       state.get("usage", []) + [usage_entry],
        }


# ── Node 6: format_answer ─────────────────────────────────────────────────────

def _calculate_cost(usage: list) -> dict:
    total_cost = 0.0
    breakdown  = []
    total_sq   = 0

    for entry in usage:
        model  = entry.get("model", "unknown")
        p_tok  = entry.get("prompt_tokens", 0)
        c_tok  = entry.get("completion_tokens", 0)
        sq     = entry.get("supabase_queries", 0)

        node_cost  = (
            p_tok * PRICING[model]["prompt"] +
            c_tok * PRICING[model]["completion"]
        ) if model in PRICING else 0.0

        node_total  = node_cost + sq * SUPABASE_COST_PER_QUERY
        total_cost += node_total
        total_sq   += sq

        breakdown.append({
            "node":              entry["node"],
            "model":             model,
            "prompt_tokens":     p_tok,
            "completion_tokens": c_tok,
            "supabase_queries":  sq,
            "cost_usd":          round(node_total, 6),
        })

    return {
        "breakdown":               breakdown,
        "total_prompt_tokens":     sum(e["prompt_tokens"] for e in usage),
        "total_completion_tokens": sum(e["completion_tokens"] for e in usage),
        "total_supabase_queries":  total_sq,
        "total_cost_usd":          round(total_cost, 6),
        "supabase_tier":           "free" if SUPABASE_COST_PER_QUERY == 0 else "pro",
    }


def _format_rows(rows: list, question_type: str) -> str:
    """
    Format Supabase rows into a readable structured block
    based on the question type.
    """
    if not rows:
        return "  No data found for this query."

    if question_type == "supplier_count":
        # Read first numeric value in the row regardless of column alias
        # (LLM may alias as 'count', 'supplier_count', 'active_supplier_count', etc.)
        row   = rows[0]
        count = next((v for v in row.values() if isinstance(v, (int, float))), "N/A")
        return f"  Total active suppliers: {count}"

    if question_type == "supplier_list":
        lines = [f"  {'Supplier':<16} {'Code':<8} {'Region'}"]
        lines.append(f"  {'-'*16} {'-'*8} {'-'*16}")
        for row in rows:
            lines.append(
                f"  {row.get('name', row.get('supplier_name', 'N/A')):<16} "
                f"{row.get('code', row.get('supplier_code', 'N/A')):<8} "
                f"{row.get('region', row.get('supplier_region', 'N/A'))}"
            )
        return "\n".join(lines)

    if question_type == "product_list":
        # Detect result shape — supplier-list vs product-list
        sample = rows[0]
        if "supplier_name" in sample or "supplier_id" in sample:
            # Supplier-list shape: supplier_id, supplier_name, supplier_code, supplier_region
            lines = [f"  {'Supplier':<16} {'Code':<8} {'Region'}"]
            lines.append(f"  {'-'*16} {'-'*8} {'-'*16}")
            for row in rows:
                lines.append(
                    f"  {row.get('supplier_name', row.get('name', 'N/A')):<16} "
                    f"{row.get('supplier_code', row.get('code', 'N/A')):<8} "
                    f"{row.get('supplier_region', row.get('region', 'N/A'))}"
                )
        else:
            # Product-list shape: rate_code, rate_type, product_type, source/destination
            lines = [f"  {'Rate Code':<12} {'Type':<8} {'Product':<14} {'Route'}"]
            lines.append(f"  {'-'*12} {'-'*8} {'-'*14} {'-'*20}")
            for row in rows:
                route = f"{row.get('source_country','?')} → {row.get('destination_country','?')}"
                lines.append(
                    f"  {row.get('rate_code','N/A'):<12} "
                    f"{row.get('rate_type','N/A'):<8} "
                    f"{row.get('product_type','N/A'):<14} "
                    f"{route}"
                )
        return "\n".join(lines)

    if question_type == "product_details":
        # Columns: rate_code, rate_type, supplier, source_country, destination_country
        # Supplier may come back as: supplier, name, supplier_name, or product_type fallback
        lines = [f"  {'Rate Code':<12} {'Type':<8} {'Supplier':<12} {'Route'}"]
        lines.append(f"  {'-'*12} {'-'*8} {'-'*12} {'-'*20}")
        for row in rows:
            route = f"{row.get('source_country','?')} → {row.get('destination_country','?')}"
            supplier = (
                row.get('supplier')
                or row.get('name')
                or row.get('supplier_name')
                or row.get('product_type')
                or 'N/A'
            )
            lines.append(
                f"  {row.get('rate_code','N/A'):<12} "
                f"{row.get('rate_type','N/A'):<8} "
                f"{supplier:<12} "
                f"{route}"
            )
        return "\n".join(lines)

    # Fallback: generic key-value dump
    lines = []
    for i, row in enumerate(rows, 1):
        lines.append(f"  [{i}] " + " | ".join(f"{k}: {v}" for k, v in row.items()))
    return "\n".join(lines)


def _write_xlsx(state: AgentState, sf: dict, rows: list) -> str:
    """
    Write a two-sheet Excel file to ./exports/:
      Sheet 1 — Client Profile: Salesforce summary card + query metadata
      Sheet 2 — Products:       clean data table, ready to filter in Excel
    Returns the file path.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)

    client_name   = (state.get("client_name") or "unknown").replace(" ", "_")
    question_type = state.get("question_type") or "query"
    today         = date.today().isoformat()
    filename      = f"{client_name}_{question_type}_{today}.xlsx"
    filepath      = os.path.join(EXPORT_DIR, filename)

    wb = openpyxl.Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────────
    DARK_BLUE   = "1F3864"
    MID_BLUE    = "2E75B6"
    LIGHT_BLUE  = "D6E4F0"
    LIGHT_GREY  = "F2F2F2"
    WHITE       = "FFFFFF"

    def header_font(bold=True, color=WHITE, size=11):
        return Font(bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)

    def center():
        return Alignment(horizontal="center", vertical="center")

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1: Client Profile ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Client Profile"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 42

    def profile_row(ws, label, value, row_idx, label_bold=False):
        lc = ws.cell(row=row_idx, column=1, value=label)
        vc = ws.cell(row=row_idx, column=2, value=value)
        lc.font   = Font(bold=label_bold, size=10, color="374151")
        vc.font   = Font(size=10, color="0F172A")
        lc.fill   = fill(LIGHT_GREY)
        vc.fill   = fill(WHITE)
        lc.border = border()
        vc.border = border()
        lc.alignment = left()
        vc.alignment = left()
        ws.row_dimensions[row_idx].height = 18

    # Title row
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value     = "KAM Supply Intelligence Agent — Client Profile"
    title_cell.font      = header_font(size=12)
    title_cell.fill      = fill(DARK_BLUE)
    title_cell.alignment = center()
    ws1.row_dimensions[1].height = 28

    # Metadata
    ws1.merge_cells("A2:B2")
    meta = ws1["A2"]
    meta.value     = f"Export date: {today}   |   Question: {state.get('question', '')}"
    meta.font      = Font(size=9, italic=True, color="6B7280")
    meta.fill      = fill(LIGHT_GREY)
    meta.alignment = left()
    ws1.row_dimensions[2].height = 16

    # Section header
    ws1.merge_cells("A3:B3")
    sh = ws1["A3"]
    sh.value     = "CLIENT PROFILE  (Salesforce)"
    sh.font      = header_font(size=10, color=WHITE)
    sh.fill      = fill(MID_BLUE)
    sh.alignment = left()
    ws1.row_dimensions[3].height = 20

    # Profile fields
    r = 4
    if sf:
        for label, key in [
            ("Client",          "Name"),
            ("Client ID",       "account_number"),
            ("Account Tier",    "account_tier"),
            ("Business Model",  "business_model"),
            ("Contract Status", "contract_status"),
            ("KAM",             "kam"),
        ]:
            profile_row(ws1, label, sf.get(key, "N/A"), r, label_bold=True)
            r += 1
    else:
        ws1.merge_cells(f"A{r}:B{r}")
        c = ws1[f"A{r}"]
        c.value = "Client profile unavailable (Salesforce error)"
        c.font  = Font(italic=True, color="DC2626", size=10)
        r += 1

    # SQL section header
    r += 1
    ws1.merge_cells(f"A{r}:B{r}")
    sq_h = ws1[f"A{r}"]
    sq_h.value     = "SQL EXECUTED"
    sq_h.font      = header_font(size=10, color=WHITE)
    sq_h.fill      = fill(MID_BLUE)
    sq_h.alignment = left()
    ws1.row_dimensions[r].height = 20
    r += 1

    sql_text = (state.get("sql_query") or "N/A").replace("\n", " ")
    ws1.merge_cells(f"A{r}:B{r}")
    sq_v = ws1[f"A{r}"]
    sq_v.value     = sql_text
    sq_v.font      = Font(size=9, color="374151", name="Courier New")
    sq_v.fill      = fill(LIGHT_GREY)
    sq_v.alignment = left()
    ws1.row_dimensions[r].height = 32
    r += 1

    # Cost footer
    r += 1
    cs = state.get("cost_summary") or {}
    ws1.merge_cells(f"A{r}:B{r}")
    cf = ws1[f"A{r}"]
    cf.value = (
        f"Query cost: ${cs.get('total_cost_usd', 0):.6f}  |  "
        f"{cs.get('total_prompt_tokens', 0)}↑ {cs.get('total_completion_tokens', 0)}↓ tokens  |  "
        f"⚠ Answers based on mock data."
    )
    cf.font      = Font(size=9, italic=True, color="6B7280")
    cf.fill      = fill(LIGHT_GREY)
    cf.alignment = left()

    # ── Sheet 2: Products ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Products")

    if rows:
        headers = list(rows[0].keys())

        # Set column widths
        col_widths = {"rate_code": 14, "rate_type": 12, "supplier": 14,
                      "product_type": 16, "source_country": 16,
                      "destination_country": 20, "count": 12}
        for i, h in enumerate(headers, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_widths.get(h, 16)

        # Header row
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h.replace("_", " ").title())
            c.font      = header_font(size=10, color=WHITE)
            c.fill      = fill(DARK_BLUE)
            c.alignment = center()
            c.border    = border()
        ws2.row_dimensions[1].height = 22

        # Data rows with alternating fill
        for row_idx, row in enumerate(rows, 2):
            row_fill = fill(LIGHT_BLUE) if row_idx % 2 == 0 else fill(WHITE)
            for col, key in enumerate(headers, 1):
                c = ws2.cell(row=row_idx, column=col, value=row.get(key, ""))
                c.font      = Font(size=10, color="0F172A")
                c.fill      = row_fill
                c.border    = border()
                c.alignment = center()
            ws2.row_dimensions[row_idx].height = 18

        # Freeze header row and enable autofilter
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

        # Client info header above the table
        ws2.insert_rows(1)
        ws2.merge_cells(f"A1:{openpyxl.utils.get_column_letter(len(headers))}1")
        ci = ws2["A1"]
        ci.value = (
            f"{sf.get('Name', 'N/A')} (ID: {sf.get('account_number', 'N/A')})  |  "
            f"{sf.get('account_tier', 'N/A')}  |  "
            f"{sf.get('business_model', 'N/A')}  |  "
            f"KAM: {sf.get('kam', 'N/A')}"
        ) if sf else "Client profile unavailable"
        ci.font      = Font(bold=True, size=10, color=WHITE)
        ci.fill      = fill(MID_BLUE)
        ci.alignment = left()
        ws2.row_dimensions[1].height = 20

    else:
        ws2["A1"] = "No data returned for this query."
        ws2["A1"].font = Font(italic=True, color="6B7280")

    wb.save(filepath)
    print(f"  → XLSX exported: {filepath}")
    return filepath


def format_answer(state: AgentState) -> AgentState:
    """
    Node 6 — Build structured answer combining Salesforce profile
    and Supabase results. Optionally export to CSV.
    """
    print(f"\n[Node 6] format_answer")

    usage        = state.get("usage", [])
    cost_summary = _calculate_cost(usage)

    print(f"  → Total cost:    ${cost_summary['total_cost_usd']:.6f}")
    print(f"  → Total tokens:  {cost_summary['total_prompt_tokens']} in / "
          f"{cost_summary['total_completion_tokens']} out")
    print(f"  → Supabase queries: {cost_summary['total_supabase_queries']} "
          f"({cost_summary['supabase_tier']} tier)")

    sf            = state.get("salesforce_data")
    rows          = state.get("sql_result") or []
    question_type = state.get("question_type", "")
    sql           = state.get("sql_query", "N/A")

    # ── Section 1: Client profile ─────────────────────────────────────────────
    if sf:
        sf_section = (
            f"┌─ CLIENT PROFILE (Salesforce) {'─' * 28}┐\n"
            f"  Client:          {sf.get('Name', 'N/A')} (ID: {sf.get('account_number', 'N/A')})\n"
            f"  Account Tier:    {sf.get('account_tier', 'N/A')}\n"
            f"  Business Model:  {sf.get('business_model', 'N/A')}\n"
            f"  Contract Status: {sf.get('contract_status', 'N/A')}\n"
            f"  KAM:             {sf.get('kam', 'N/A')}\n"
            f"└{'─' * 59}┘"
        )
    else:
        sf_section = (
            f"⚠️  CLIENT PROFILE UNAVAILABLE\n"
            f"  Salesforce error: {state.get('salesforce_error', 'unknown')}\n"
            f"  Showing operational data only."
        )

    # ── Section 2: Operational data ───────────────────────────────────────────
    if state.get("sql_error"):
        data_section = (
            f"┌─ OPERATIONAL DATA (Supabase) {'─' * 29}┐\n"
            f"  ⚠️  Query failed after {state.get('retry_count', 0)} retries.\n"
            f"  Error: {str(state.get('sql_error', 'Unknown error'))[:120]}\n"
            f"  Please rephrase your question or contact support.\n"
            f"└{'─' * 59}┘"
        )
    else:
        data_section = (
            f"┌─ OPERATIONAL DATA (Supabase) {'─' * 29}┐\n"
            f"{_format_rows(rows, question_type)}\n"
            f"└{'─' * 59}┘"
        )

    # ── Section 3: Transparency ───────────────────────────────────────────────
    sql_section = (
        f"┌─ SQL EXECUTED {'─' * 44}┐\n"
        f"  {sql}\n"
        f"└{'─' * 59}┘"
    )

    # ── Cost footer ───────────────────────────────────────────────────────────
    cost_line = (
        f"💰 Query cost: ${cost_summary['total_cost_usd']:.6f} | "
        f"{cost_summary['total_prompt_tokens']}↑ "
        f"{cost_summary['total_completion_tokens']}↓ tokens | "
        f"{cost_summary['total_supabase_queries']} Supabase "
        f"{'query' if cost_summary['total_supabase_queries'] == 1 else 'queries'} "
        f"({cost_summary['supabase_tier']} tier)\n"
        f"⚠️  Answers based on mock data."
    )

    answer = f"{sf_section}\n\n{data_section}\n\n{sql_section}\n\n{cost_line}"

    # ── Optional CSV export ───────────────────────────────────────────────────
    csv_path = None
    if state.get("export_csv") and rows:
        csv_path = _write_xlsx(state, sf, rows)

    return {
        **state,
        "usage":        usage,
        "cost_summary": cost_summary,
        "final_answer": answer,
        "csv_path":     csv_path,
    }


# ── Routing ───────────────────────────────────────────────────────────────────

MAX_RETRIES = 2

def route_after_understand_question(state: AgentState) -> str:
    if not state.get("client_name"):
        print(f"  → No client name — short-circuiting to END")
        return END
    return "fetch_salesforce_client"

def route_after_execute_sql(state: AgentState) -> str:
    if state.get("sql_error") and state.get("retry_count", 0) < MAX_RETRIES:
        print(f"  → SQL error, routing back to generate_sql "
              f"(retry {state['retry_count']})")
        return "generate_sql"
    return "format_answer"


# ── Build graph ───────────────────────────────────────────────────────────────

def build_graph() -> StateGraph:
    graph = StateGraph(AgentState)

    graph.add_node("understand_question",     understand_question)
    graph.add_node("fetch_salesforce_client", fetch_salesforce_client)
    graph.add_node("retrieve_schema",         retrieve_schema)
    graph.add_node("generate_sql",            generate_sql)
    graph.add_node("execute_sql",             execute_sql)
    graph.add_node("format_answer",           format_answer)

    graph.set_entry_point("understand_question")

    graph.add_conditional_edges(
        "understand_question",
        route_after_understand_question,
        {"fetch_salesforce_client": "fetch_salesforce_client", END: END},
    )
    graph.add_edge("fetch_salesforce_client", "retrieve_schema")
    graph.add_edge("retrieve_schema",         "generate_sql")
    graph.add_edge("generate_sql",            "execute_sql")
    graph.add_conditional_edges(
        "execute_sql",
        route_after_execute_sql,
        {"generate_sql": "generate_sql", "format_answer": "format_answer"},
    )
    graph.add_edge("format_answer", END)

    return graph.compile()


# ── Public entrypoint ─────────────────────────────────────────────────────────

def run_agent(question: str, export_csv: bool = False) -> dict:
    app = build_graph()
    initial_state: AgentState = {
        "question":          question,
        "export_csv":        export_csv,
        "client_name":       None,
        "question_type":     None,
        "intent_summary":    None,
        "salesforce_data":   None,
        "salesforce_error":  None,
        "schema_context":    None,
        "schema_doc_ids":    None,
        "sql_query":         None,
        "sql_result":        None,
        "sql_error":         None,
        "retry_count":       0,
        "final_answer":      None,
        "csv_path":          None,   # .xlsx path if exported
        "usage":             [],
        "cost_summary":      None,
    }
    return app.invoke(initial_state)


# ── Smoke test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    test_questions = [
        # Run with export_csv=True on Q3 to test CSV export
        ("How many suppliers does Check24 have?",                          False),
        ("Which products does Avis have connected to Autoslash?",          False),
        ("What are the details of Check24's inbound products from Germany?", True),
        ("How many suppliers does Booking.com have?",                      False),
    ]

    for q, export in test_questions:
        print("\n" + "=" * 60)
        result = run_agent(q, export_csv=export)
        print(f"\n── FINAL ANSWER ──")
        print(result["final_answer"])
        if result.get("csv_path"):
            print(f"\n── CSV EXPORTED ──")
            print(f"  {result['csv_path']}")
        print(f"\n── SCHEMA RETRIEVED ──")
        for doc_id in (result.get("schema_doc_ids") or []):
            print(f"  {doc_id}")
        print(f"\n── PARSED INTENT ──")
        print(f"  client_name:   {result['client_name']}")
        print(f"  question_type: {result['question_type']}")
        print(f"  intent:        {result['intent_summary']}")
        if result.get("cost_summary"):
            print(f"\n── COST SUMMARY ──")
            cs = result["cost_summary"]
            for row in cs.get("breakdown", []):
                print(f"  {row['node']:<30} {row['model']:<26} "
                      f"{row['prompt_tokens']:>5}↑ {row['completion_tokens']:>4}↓ tokens  "
                      f"sq:{row['supabase_queries']}  ${row['cost_usd']:.6f}")
            print(f"  {'TOTAL':<30} {'':26} "
                  f"{cs.get('total_prompt_tokens',0):>5}↑ "
                  f"{cs.get('total_completion_tokens',0):>4}↓ tokens  "
                  f"sq:{cs.get('total_supabase_queries',0)}  "
                  f"${cs.get('total_cost_usd',0):.6f}")
